from time import sleep
import pyautogui
from PIL import ImageGrab, Image
import pyscreeze
import cv2
import random
import time
import os
import string
import random
import string
import openpyxl
from actions import *

file_path = 'steps.xlsx'
workbook=openpyxl.reader.excel.load_workbook(file_path)
sheet = workbook.active
start_row=2
start_column=1
skip_counter=0

print(len(sheet['A']))
    
def takeScreenShot():
    """
    使用 pyscreeze 库截取屏幕截图，并返回读取的图片

    返回:
        numpy.ndarray: 读取的屏幕截图图片
    """
    # 截取屏幕截图并保存为 my_screenshot.png 文件
    screenshot = pyscreeze.screenshot('screenshot.png')
    
    # 读取保存的截图文件，以灰度模式读取
    screenshot = cv2.imread('screenshot.png')
    
    # 释放内存，删除临时截图文件
    os.remove('screenshot.png')
    
    # 返回读取的屏幕截图图片
    return screenshot

def getTarget(filePath):
    """
    根据传入的步数获取对应的目标图片

    参数:
        step (int): 步数

    返回:
        numpy.ndarray: 读取的目标图片
    """
    # 检查文件是否存在
    if os.path.exists(filePath):
        # 事先读取按钮截图
        target = cv2.imread(filePath, cv2.IMREAD_GRAYSCALE)
        return target
    else:
        raise Exception("文件不存在")

def excute(currentRow):    
    global skip_counter
    
    action=sheet.cell(row=currentRow, column=start_column+1).value
    if action=="Wait":
        time.sleep(int(sheet.cell(row=currentRow, column=start_column+5).value))
        return True
    
    targetPath=sheet.cell(row=currentRow, column=start_column+2).value
    target=getTarget(targetPath)
    
    screenShot=takeScreenShot() 
    if action=="Click":
        while not clickAction(screenShot,target):
            time.sleep(2)
            screenShot=takeScreenShot()
        return True
    elif action=="Move":
        while not moveAction(screenShot,target):
            time.sleep(2)
            screenShot=takeScreenShot() 
        return True
    elif action=="Scroll":
        scroll=sheet.cell(row=currentRow, column=start_column+6).value
        while not scrollAction(screenShot, scroll, target):
            time.sleep(2)
            screenShot=takeScreenShot() 
        return True
    elif action=="Type":
        text=sheet.cell(row=currentRow, column=start_column+4).value
        while not scrollAction(screenShot, text, target):
            time.sleep(2)
            screenShot=takeScreenShot() 
        return True
    elif action=="Drag":
        target2=getTarget(sheet.cell(row=currentRow, column=start_column+8).value)
        while not dragAction(screenShot,target,target2):
            time.sleep(2)
            screenShot=takeScreenShot() 
        return True
    elif action=="Right Click":
        while not rightClickAction(screenShot,target):
            time.sleep(2)
            screenShot=takeScreenShot()
        return True
    elif action=="Jump":
        if checkExistAction(screenShot,target):
            skip_counter=int(sheet.cell(row=currentRow, column=start_column+3).value)
            return True
    elif action=="Hold":
        while not holdAction(screenShot,int(sheet.cell(row=currentRow, column=start_column+5).value),target):
            time.sleep(2)
            screenShot=takeScreenShot()
        return True
    else:
        print("没有找到对应的操作")
        
if __name__ == '__main__':
    i=len(sheet['A'])-1
    
    n=0
    while n<=i:
        if skip_counter>0:
            n=skip_counter
        print("第"+str(n)+"行")
        excute(n)
    print("完成")